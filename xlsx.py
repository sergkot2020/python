# -*- coding: utf-8 -*-
import openpyxl
import os
import copy
import csv

all_file = os.listdir()
all_program = []
all_computers = []
clear_list = []
ignor_list = []

#  функция загркузки xlsx файла
def load_xlsx(file):
    xls = file.split('.')[-1]
    if xls == 'xlsx':
        wb = openpyxl.load_workbook(file)
        sheet = wb['Лист1']
        return sheet


# ищем все программы по всех файлах и составляем список всех программ
for file in all_file:
    sheet = load_xlsx(file)
    if sheet is not None:
        for i in range(1, 200):
            prog_name = sheet.cell(row=i, column=1).value
            if prog_name is None:
                break
            if not prog_name in all_program:
                all_program.append(prog_name)

# пишем список всех программ в текстовый файл
if os.path.isfile('all_prog.txt') is not True:
    with open('all_prog.txt', 'a', encoding='utf-8') as f_obj:
        for proga in all_program:
            string = f'{proga}\n'
            f_obj.write(string)

# Подгружаем в память список игнорируемых программ
with open('ignor_list.txt', 'r', encoding='utf-8') as ignor:
    for proga in ignor:
        proga = proga.split('\n')[0]
        ignor_list.append(proga)

# чистим список с программами
for ignor in ignor_list:
    if ignor in all_program:
        index = all_program.index(ignor)
        all_program.pop(index)


# сортируем список по алфавиту
all_program.sort()
all_program.insert(0, 'Комп')

# генерис список, который потом будем заоплнять программами в будущей таблице
clear_list = [x*0 for x in range(len(all_program))]


# сверяем программы из каждого файла со списком всех программ
for file in all_file:
    sheet = load_xlsx(file)
    if sheet is not None:
        buffer = copy.deepcopy(clear_list)
        for i in range(1, 200):
            prog_name = sheet.cell(row=i, column=1).value
            if prog_name is None:
                break
            for index, count in enumerate(all_program):
                if prog_name == count or buffer[index] == 1:
                    buffer[index] = 1
                else:
                    buffer[index] = 0
        buffer[0] = file
        all_computers.append(buffer)
        buffer = clear_list

# записываем результат в отчет в формате csv
result = "result.csv"
data = [[1, 1, 1], [2,2, 2], [3, 3, 3]]
with open(result, "w", newline='') as csv_file:
    writer = csv.writer(csv_file, delimiter=',')
    for index, comp in enumerate(all_computers):
        if index == 0:
            writer.writerow(all_program)
        writer.writerow(comp)


